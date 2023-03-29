import tkinter as tk
import openpyxl
from openpyxl import Workbook
from tkinter import messagebox
from tkinter import *
from tkinter import ttk


# Modelo
class Demanda:
    def __init__(self, data, numero, setor, assunto, descricao, status):
        self.__data = data
        self.__numero = numero
        self.__setor = setor
        self.__assunto = assunto
        self.__descricao = descricao
        self.__status = status

    def get_data(self):
        return self.__data

    def set_data(self, data):
        self.__data = data

    def get_numero(self):
        return self.__numero

    def set_numero(self, numero):
        self.__numero = numero

    def get_setor(self):
        return self.__setor

    def set_setor(self, setor):
        self.__setor = setor

    def get_assunto(self):
        return self.__assunto

    def set_assunto(self, assunto):
        self.__assunto = assunto

    def get_descricao(self):
        return self.__descricao

    def set_descricao(self, descricao):
        self.__descricao = descricao

    def get_status(self):
        return self.__status

    def set_status(self, status):
        self.__status = status


class Modelo:
    def __init__(self):
        self.demandas = []
        self.ultimo_numero = 0
        self.arquivo_excel = 'demandas.xlsx'

        try:
            self.workbook = openpyxl.load_workbook(self.arquivo_excel)
            self.worksheet = self.workbook.active
        except FileNotFoundError:
            self.workbook = Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.append(['Data de Recebimento', 'Nº do memorando', 'Setor/Escola', 'Assunto', 'Descrição', 'Status'])
            self.workbook.save(self.arquivo_excel)

    def criar_demanda(self, data, setor, assunto, descricao, status):
        self.ultimo_numero += 1
        numero = self.ultimo_numero
        demanda = Demanda(data, numero, setor, assunto, descricao, status)
        self.demandas.append(demanda)
        self.worksheet.append([data, numero, setor, assunto, descricao, status])
        self.workbook.save(self.arquivo_excel)


    def atualizar_demanda(self, numero, **kwargs):
        demanda = self.ler_demanda(numero)
        if demanda is not None:
            for key, value in kwargs.items():
                setattr(demanda, key, value)

            for row in self.worksheet.iter_rows(min_row=2, min_col=1, max_col=6):
                if row[1].value == numero:
                    row[0].value = demanda.get_data()
                    row[2].value = demanda.get_setor()
                    row[3].value = demanda.get_assunto()
                    row[4].value = demanda.get_descricao()
                    row[5].value = demanda.get_status()

            self.workbook.save(self.arquivo_excel)

    def excluir_demanda(self, numero):
        demanda = self.ler_demanda(numero)
        if demand
